#!/usr/bin/env python3
"""
SH철강 운영 데이터 마이그레이션 도구

xlsx와 CSV 사이의 양방향 동기화 + 자동 미러링.
xlsx는 view, CSV가 truth source.

usage:
    python migrate.py extract                      # xlsx → CSV (데이터 추출)
    python migrate.py rebuild                      # 빈 양식 + CSV → xlsx
    python migrate.py mirror                       # 매출 수금완료 → 통장 자동 미러링
    python migrate.py sync                         # extract + mirror (일상 운영)
    python migrate.py refresh                      # extract + mirror + rebuild (양식 변경 사이클)
    python migrate.py extract --xlsx path/to.xlsx  # 다른 xlsx 지정

명령 의미:
    sync     — xlsx 편집 후 백업 + 미러링. xlsx는 그대로 유지 (Excel 캐시 보존).
    refresh  — 양식(create_spreadsheets.py)이 바뀐 후 xlsx까지 재생성.
    mirror   — 매출 수금완료 → 6.통장 행 자동 INSERT (매칭ID dedup).
"""

import csv
import sys
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook

# 동일 패키지의 create_spreadsheets에서 build_workbook 가져오기
sys.path.insert(0, str(Path(__file__).parent / 'workbook'))
from create_spreadsheets import build_workbook  # noqa: E402


# ============================================================
# 시트별 메타데이터
# ============================================================
# header_row: 헤더가 있는 행 (데이터는 그 다음 행부터)
# max_col: 데이터 영역의 최대 컬럼 (영수증 시트는 A~H만, J~L은 합계 영역)
# formula_cols: 자동 수식 컬럼 (1-based) — CSV에 저장하지 않고, rebuild 시 건드리지 않음
# date_cols: 날짜 컬럼 (1-based) — CSV ↔ xlsx 변환 시 datetime 처리
# skip_label_rows: 1열 값이 이 set에 속하는 행은 데이터가 아닌 양식의 일부로 간주하고 추출 제외
#                  (예: 8.미수관리의 row 57 "합계" 행이 데이터로 추출되는 것을 방지)
SHEETS = {
    '1.매출': {
        'header_row': 1,
        'max_col': 22,
        'formula_cols': [2, 3, 15, 20, 21],   # 주문ID, 상태, 수금예정일, 미수금, 미수일수
        'date_cols': [1, 13],                  # 기록일, 납품일자 (수금예정일 O는 수식이라 제외)
    },
    '2.매입': {
        'header_row': 1,
        'max_col': 21,
        'formula_cols': [2, 3, 14, 19, 20],   # 매입ID, 상태, 결제예정일, 미지급금, 지연일수
        'date_cols': [1, 12],                  # 기록일, 입고일자
    },
    '3.재고': {
        'header_row': 1,
        'max_col': 12,
        'formula_cols': [8, 11],               # 재고가치, 회전상태
        'date_cols': [9, 10],                  # 마지막입고일, 마지막출고일
    },
    '4.영수증': {
        'header_row': 4,
        'max_col': 8,                          # A~H만 데이터 (J~L은 합계 영역)
        'formula_cols': [2],                   # 영수증ID
        'date_cols': [1],                      # 날짜
    },
    '5.거래처': {
        'header_row': 1,
        'max_col': 12,
        'formula_cols': [10, 11],              # 미수금, 누적매출
        'date_cols': [7],                      # 거래시작일
    },
    '6.통장': {
        'header_row': 1,
        'max_col': 10,                         # A~J만 (L~O는 잔고 요약 영역)
        'formula_cols': [7],                   # 누적잔고
        'date_cols': [1],                      # 일자
    },
    '7.정기업무': {
        'header_row': 4,
        'max_col': 9,
        'formula_cols': [],
        'date_cols': [4, 5],                   # 마지막 실행일, 다음 실행일
    },
    '8.미수관리': {
        'header_row': 4,
        'max_col': 10,
        'formula_cols': [2, 3, 4, 5, 6, 7, 8],  # 모든 집계는 수식
        'date_cols': [9],                       # 마지막 독촉일
        'skip_label_rows': {'합계'},            # row 57의 합계 라벨 행 (양식 일부) 제외
    },
    '0.개선아이디어': {
        'header_row': 4,
        'max_col': 6,
        'formula_cols': [],
        'date_cols': [1],
    },
    '9.영업내역': {
        'header_row': 1,
        'max_col': 15,
        'formula_cols': [2, 4],                # 활동ID, 등록여부
        'date_cols': [1, 14],                  # 일자, 다음 follow-up
    },
}

# 기본 경로
DEFAULT_XLSX = 'workbook/output/SL철강_5월_운영시트.xlsx'
DATA_DIR = Path('data')


# ============================================================
# 값 변환 헬퍼
# ============================================================
def to_csv_value(value, is_date_col):
    """xlsx 값 → CSV에 쓸 문자열"""
    if value is None:
        return ''
    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d')
    return str(value)


def from_csv_value(s, is_date_col):
    """CSV 문자열 → xlsx에 넣을 값"""
    if s == '' or s is None:
        return None
    if is_date_col:
        # 날짜 파싱 시도
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except ValueError:
            pass
        try:
            return datetime.strptime(s, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            pass
        return s  # 파싱 실패 시 문자열 그대로
    # 숫자 시도
    try:
        if '.' in s:
            return float(s)
        return int(s)
    except ValueError:
        pass
    return s


def get_headers(ws, header_row, max_col):
    """헤더 행에서 컬럼명 추출"""
    return [ws.cell(row=header_row, column=c).value for c in range(1, max_col + 1)]


# ============================================================
# extract: xlsx → CSV
# ============================================================
def extract(xlsx_path: str = DEFAULT_XLSX, data_dir: Path = DATA_DIR):
    """현재 xlsx의 입력 데이터를 CSV로 추출 (자동수식 컬럼은 제외)"""
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        print(f'✗ xlsx not found: {xlsx_path}')
        sys.exit(1)

    wb = load_workbook(xlsx_path, data_only=True)
    data_dir.mkdir(parents=True, exist_ok=True)

    total_rows = 0
    for sheet_name, meta in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f'⚠ skip {sheet_name} (시트 없음)')
            continue
        ws = wb[sheet_name]
        header_row = meta['header_row']
        max_col = meta['max_col']
        date_cols = set(meta['date_cols'])
        skip_labels = set(meta.get('skip_label_rows', ()))

        headers = get_headers(ws, header_row, max_col)
        rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            row_values = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            # 첫 번째 데이터 컬럼이 비어있으면 빈 행으로 간주
            # 단, 매출/매입은 1열이 기록일이라 그 기준
            # 거래처/재고/통장/영수증/미수관리/개선아이디어/정기업무도 1열 기준 OK
            if row_values[0] is None or row_values[0] == '':
                continue
            # 양식의 라벨 행 (합계, 소계 등) 제외 — 시트별 skip_label_rows 설정
            if row_values[0] in skip_labels:
                continue
            rows.append(row_values)

        csv_path = data_dir / f'{sheet_name}.csv'
        with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row in rows:
                csv_row = []
                for col_idx, value in enumerate(row, 1):
                    is_date = col_idx in date_cols
                    csv_row.append(to_csv_value(value, is_date))
                writer.writerow(csv_row)

        print(f'✓ {sheet_name:18} → {csv_path.name:25} ({len(rows)}행)')
        total_rows += len(rows)

    print(f'\n총 {total_rows}행 추출 완료. data/ 디렉토리 확인.')


# ============================================================
# rebuild: 빈 양식 + CSV → xlsx
# ============================================================
def rebuild(xlsx_path: str = DEFAULT_XLSX, data_dir: Path = DATA_DIR):
    """빈 양식 새로 생성하고 CSV 데이터를 주입"""
    xlsx_path = Path(xlsx_path)

    # 1. 빈 양식 생성 (with_samples=False)
    print(f'1) 빈 양식 생성: {xlsx_path}')
    build_workbook(str(xlsx_path), with_samples=False)

    # 2. CSV 데이터 주입
    if not data_dir.exists():
        print(f'⚠ {data_dir}/ 없음 — 빈 양식 그대로 저장')
        return

    wb = load_workbook(xlsx_path)
    total_rows = 0

    for sheet_name, meta in SHEETS.items():
        csv_path = data_dir / f'{sheet_name}.csv'
        if not csv_path.exists():
            print(f'⚠ skip {sheet_name} (CSV 없음)')
            continue

        ws = wb[sheet_name]
        header_row = meta['header_row']
        max_col = meta['max_col']
        formula_cols = set(meta['formula_cols'])
        date_cols = set(meta['date_cols'])

        with open(csv_path, encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            csv_headers = next(reader)  # 헤더 스킵
            row_count = 0
            for i, row in enumerate(reader):
                target_row = header_row + 1 + i
                for col_idx, value_str in enumerate(row, 1):
                    if col_idx > max_col:
                        break
                    if col_idx in formula_cols:
                        continue  # 수식 컬럼은 건드리지 않음
                    is_date = col_idx in date_cols
                    parsed = from_csv_value(value_str, is_date)
                    if parsed is not None:
                        ws.cell(row=target_row, column=col_idx, value=parsed)
                row_count += 1

        print(f'✓ {sheet_name:18} ← {csv_path.name:25} ({row_count}행)')
        total_rows += row_count

    wb.save(xlsx_path)
    print(f'\n총 {total_rows}행 주입 완료. {xlsx_path} 저장됨.')


# ============================================================
# mirror: 매출 수금완료 → 6.통장 자동 행 추가
# ============================================================
# 매출에서 (금액 ✓ + 수금완료=O + 입금통장 ✓) 행을 찾아 6.통장에 입금 행으로 INSERT.
# 6.통장의 매칭ID로 dedup — 이미 미러링된 행은 skip (idempotent).
# 사용자가 6.통장 행을 수동 편집하면 그 변경은 보존됨 (mirror가 update X, insert만).
# ============================================================
def mirror_sales_to_bank(data_dir: Path = DATA_DIR):
    """매출 수금완료 → 6.통장 자동 미러링 (INSERT-only, 매칭ID dedup)"""
    sales_csv = data_dir / '1.매출.csv'
    bank_csv = data_dir / '6.통장.csv'

    if not sales_csv.exists():
        print(f'⚠ {sales_csv} 없음 — mirror skip')
        return 0
    if not bank_csv.exists():
        print(f'⚠ {bank_csv} 없음 — mirror skip')
        return 0

    # 매출 읽기
    with open(sales_csv, encoding='utf-8-sig') as f:
        sales_rows = list(csv.DictReader(f))

    # 통장 읽기 (헤더 + 행)
    with open(bank_csv, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        bank_lines = list(reader)
    bank_header = bank_lines[0]
    bank_data = bank_lines[1:]

    # 시작잔고 행과 거래 행 분리 (분류='시작' 기준)
    # 6.통장 컬럼: 0=일자, 1=사업자, 2=통장, 3=적요, 4=입금, 5=출금,
    #              6=누적잔고(수식), 7=분류, 8=매칭ID, 9=메모
    starting_rows = []
    transaction_rows = []
    for row in bank_data:
        category = row[7] if len(row) > 7 else ''
        if category == '시작':
            starting_rows.append(row)
        else:
            transaction_rows.append(row)

    # 기존 매칭ID 수집 (dedup용)
    existing_match_ids = set()
    for row in transaction_rows:
        match_id = row[8] if len(row) > 8 else ''
        if match_id:
            existing_match_ids.add(match_id)

    # 매출에서 미러링 대상 추출 + 새 통장 행 생성
    added = 0
    skipped_no_account = []
    for s in sales_rows:
        # 금액 0 또는 비어있으면 skip
        amount_str = s.get('금액(원)', '').strip()
        try:
            amount = float(amount_str) if amount_str else 0
        except ValueError:
            continue
        if amount == 0:
            continue
        # 수금완료=O 아니면 skip
        if s.get('수금완료', '').strip() != 'O':
            continue
        # 주문ID 없으면 skip (수식 미평가 상태 — 사용자가 xlsx 한 번 열면 채워짐)
        order_id = s.get('주문ID', '').strip()
        if not order_id:
            continue
        # 이미 미러링됐으면 skip
        if order_id in existing_match_ids:
            continue
        # 입금통장 없으면 skip + 추적
        bank_account = s.get('입금통장', '').strip()
        if not bank_account:
            skipped_no_account.append((order_id, s.get('거래처', '')))
            continue

        # 통장 행 생성
        date = (s.get('납품일자') or s.get('수금예정일') or s.get('기록일') or '').strip()
        entity_raw = s.get('사업자', '').strip()
        # 사업자 차원 정규화 — 6.통장의 '사업자' 컬럼은 (법인, 사업자) 2분할.
        # 1.매출에서 사업자=B계좌로 입력된 행은 통장 차원의 'B계좌'(히든 통장)이고
        # entity 차원으로는 '사업자' 산하. 도메인_룰.md §1 참조.
        entity = '사업자' if entity_raw == 'B계좌' else entity_raw
        party = s.get('거래처', '').strip()
        site = s.get('현장', '').strip()
        memo = s.get('메모', '').strip()

        # 적요 — "거래처 (현장) 매출 입금"
        desc = party
        if site:
            desc += f' ({site})'
        desc += ' 매출 입금'

        # 분류 — B계좌면 별도 표시 (부가세 신고 자료 분리용)
        category = '매출입금(B통장)' if bank_account == 'B계좌' else '매출입금'

        # 입금 금액 (정수면 정수로 표기)
        amount_str_out = str(int(amount)) if amount == int(amount) else str(amount)

        new_row = [
            date,             # 일자
            entity,           # 사업자
            bank_account,     # 통장
            desc,             # 적요
            amount_str_out,   # 입금
            '0',              # 출금
            '',               # 누적잔고 (수식 컬럼 — formula_cols=[7]이라 비워둠)
            category,         # 분류
            order_id,         # 매칭ID
            memo,             # 메모
        ]
        transaction_rows.append(new_row)
        existing_match_ids.add(order_id)
        added += 1

    # skip 알림
    if skipped_no_account:
        print(f'⚠ 입금통장 미입력 매출 {len(skipped_no_account)}건 skip:')
        for oid, party in skipped_no_account:
            print(f'    {oid} — {party}')

    if added == 0:
        print(f'✓ 매출 미러링: 추가 0건 (이미 모두 미러링됨)')
        return 0

    # 거래 행 정렬: 일자 → 매칭ID 보조키
    transaction_rows.sort(key=lambda r: ((r[0] or ''), (r[8] or '')))

    # 쓰기 — 시작잔고 그대로, 거래는 정렬된 순서로
    with open(bank_csv, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(bank_header)
        for row in starting_rows:
            writer.writerow(row)
        for row in transaction_rows:
            writer.writerow(row)

    print(f'✓ 매출 미러링: {added}건 → {bank_csv.name}')
    return added


# ============================================================
# sync: extract + mirror (xlsx → CSV + 매출 미러링)
# ============================================================
# 일상 운영 명령. xlsx 편집 후 백업/미러링까지 한 번에.
# 양식 변경(create_spreadsheets.py 수정) 후 xlsx 재생성하려면 'refresh' 사용.
# ============================================================
def sync(xlsx_path: str = DEFAULT_XLSX, data_dir: Path = DATA_DIR):
    """xlsx → CSV + 매출 미러링 (xlsx는 그대로 유지)"""
    print('=' * 60)
    print('SYNC: 데이터 추출 + 매출 → 통장 미러링')
    print('=' * 60)
    print('\n[1/2] EXTRACT (xlsx → CSV)')
    print('-' * 60)
    extract(xlsx_path, data_dir)
    print('\n[2/2] MIRROR (매출 수금완료 → 통장)')
    print('-' * 60)
    mirror_sales_to_bank(data_dir)
    print('\n✓ SYNC 완료')


# ============================================================
# refresh: extract + mirror + rebuild (양식 변경 사이클)
# ============================================================
# create_spreadsheets.py의 양식이 바뀐 후 xlsx까지 재생성할 때 사용.
# 일상 sync와 분리 — Excel이 박은 수식 캐시가 rebuild로 사라지는 부작용 회피.
# ============================================================
def refresh(xlsx_path: str = DEFAULT_XLSX, data_dir: Path = DATA_DIR):
    """xlsx → CSV → 미러링 → CSV → xlsx (양식 변경 후 사이클)"""
    print('=' * 60)
    print('REFRESH: 데이터 추출 + 미러링 + 양식 재빌드')
    print('=' * 60)
    print('\n[1/3] EXTRACT')
    print('-' * 60)
    extract(xlsx_path, data_dir)
    print('\n[2/3] MIRROR')
    print('-' * 60)
    mirror_sales_to_bank(data_dir)
    print('\n[3/3] REBUILD')
    print('-' * 60)
    rebuild(xlsx_path, data_dir)
    print('\n✓ REFRESH 완료')


# ============================================================
# 진입점
# ============================================================
def parse_args():
    args = sys.argv[1:]
    cmd = 'sync'
    xlsx_path = DEFAULT_XLSX
    data_dir = DATA_DIR
    i = 0
    while i < len(args):
        a = args[i]
        if a in ('extract', 'rebuild', 'sync', 'mirror', 'refresh'):
            cmd = a
        elif a == '--xlsx':
            xlsx_path = args[i + 1]
            i += 1
        elif a == '--data':
            data_dir = Path(args[i + 1])
            i += 1
        elif a in ('-h', '--help'):
            print(__doc__)
            sys.exit(0)
        i += 1
    return cmd, xlsx_path, data_dir


if __name__ == '__main__':
    cmd, xlsx_path, data_dir = parse_args()
    if cmd == 'extract':
        extract(xlsx_path, data_dir)
    elif cmd == 'rebuild':
        rebuild(xlsx_path, data_dir)
    elif cmd == 'mirror':
        mirror_sales_to_bank(data_dir)
    elif cmd == 'sync':
        sync(xlsx_path, data_dir)
    elif cmd == 'refresh':
        refresh(xlsx_path, data_dir)
