#!/usr/bin/env python3
"""
SH철강 운영 스프레드시트 생성기 (양식 정의)

usage:
    python create_spreadsheets.py                              # samples 포함, 기본 경로
    python create_spreadsheets.py --no-samples                 # 빈 양식 (마이그레이션용)
    python create_spreadsheets.py path/to/output.xlsx          # 경로 지정
    python create_spreadsheets.py path/to/output.xlsx --no-samples
"""

from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================
# 공통 스타일 — sl-steel.co.kr 디자인 시스템과 일관
# ============================================================
FONT_NAME = '맑은 고딕'
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', start_color='2C5F8A')
INPUT_FONT = Font(name=FONT_NAME, size=10, color='0000FF')      # 파랑 = 사용자 입력
FORMULA_FONT = Font(name=FONT_NAME, size=10, color='000000')    # 검정 = 자동 수식
DEFAULT_FONT = Font(name=FONT_NAME, size=10)
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color='2C5F8A')
SUBTITLE_FONT = Font(name=FONT_NAME, size=11, bold=True, color='1E1C18')

THIN = Side(border_style='thin', color='CCCCCC')
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RED_FILL = PatternFill('solid', start_color='FFCCCC')
YELLOW_FILL = PatternFill('solid', start_color='FFFFCC')
GREEN_FILL = PatternFill('solid', start_color='CCFFCC')
LIGHT_BLUE_FILL = PatternFill('solid', start_color='E8F0F7')


# ============================================================
# 헬퍼
# ============================================================
def style_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 32


def set_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ============================================================
# 메인 빌드 함수
# ============================================================
def build_workbook(output: str = 'SL철강_5월_운영시트.xlsx', with_samples: bool = True):
    """
    워크북 생성.

    Args:
        output: 저장 경로
        with_samples: True면 샘플/프리셋 데이터 포함, False면 빈 양식 (마이그레이션용)
    """
    wb = Workbook()
    wb.remove(wb.active)

    _build_sheet_sales(wb, with_samples)
    _build_sheet_purchases(wb, with_samples)
    _build_sheet_inventory(wb, with_samples)
    _build_sheet_customers(wb, with_samples)      # 5.거래처
    _build_sheet_bank(wb, with_samples)           # 6.통장
    _build_sheet_dashboard(wb)                    # 현황판 (수식만, 데이터 없음)
    _build_sheet_ideas(wb, with_samples)          # 0.개선아이디어
    _build_sheet_recurring_tasks(wb, with_samples)  # 7.정기업무
    _build_sheet_receivables(wb, with_samples)    # 8.미수관리
    _build_sheet_receipts(wb, with_samples)       # 4.영수증

    # 시트 순서 조정 — 4.영수증을 6번째 위치(인덱스 5)로
    # 순서: 현황판(0), 0.개선(1), 1.매출(2), 2.매입(3), 3.재고(4), 4.영수증(5),
    #      5.거래처(6), 6.통장(7), 7.정기업무(8), 8.미수관리(9)
    _idx = wb.sheetnames.index('4.영수증')
    wb.move_sheet('4.영수증', offset=5 - _idx)

    # 출력 디렉토리 보장
    Path(output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return wb


# ============================================================
# 시트 1: 매출
# ============================================================
def _build_sheet_sales(wb, with_samples):
    ws = wb.create_sheet('1.매출')

    headers = [
        '기록일', '주문ID', '상태',
        '거래처', '품목', '규격', '단위', '수량', '금액(원)',
        '결제방식', '사업자',
        '납품일자', '거래명세서',
        '수금예정일', '수금완료', '세금계산서', '사진전송',
        '입금통장',
        '미수금(원)', '미수일수', '메모'
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    if with_samples:
        samples = [
            [date(2026, 5, 3), None, None,
             'OO건설', '철근', 'D13', '톤', 2.5, 1250000,
             '외상30일', '법인',
             None, '',
             None, '', '', '',
             '법인A',
             None, None, '주문 단계'],
            [date(2026, 5, 1), None, None,
             '박사장', '각파이프', '50×50×2.0t', '개', 200, 600000,
             '외상7일', '법인',
             date(2026, 5, 4), 'O',
             None, '', '', '',
             '법인A',
             None, None, '납품완료, 수금대기'],
            [date(2026, 5, 4), None, None,
             '소형시공', '철근', 'D10', '톤', 0.8, 384000,
             '현금', '사업자',
             date(2026, 5, 5), 'O',
             None, 'O', 'O', 'O',
             '사업자A',
             None, None, '수금완료'],
            [date(2026, 5, 6), None, None,
             '김씨건축', '강관', '50각', '개', 50, 250000,
             '현금', '사업자',
             date(2026, 5, 6), '',
             None, '', '', '',
             '사업자B',
             None, None, '무자료 거래, B통장 입금'],
        ]
        for s in samples:
            ws.append(s)

    # 수식 (100행까지) — 데이터 유무 상관없이 항상 채움
    for row in range(2, 102):
        ws.cell(row=row, column=2,
                value=f'=IF($A{row}="","",TEXT($A{row},"YYYYMMDD")&"-"&TEXT(COUNTIF($A$2:$A{row},$A{row}),"000"))')
        ws.cell(row=row, column=3,
                value=f'=IF($D{row}="","",IF($O{row}="O","수금완료",IF(OR($M{row}="O",$P{row}="O",$R{row}="사업자B"),IF(AND($N{row}<>"",$N{row}<TODAY()),"연체","납품완료"),"주문")))')
        ws.cell(row=row, column=14,
                value=f'=IF(L{row}="","",IF(J{row}="즉시",L{row},IF(J{row}="현금",L{row},IF(J{row}="외상1일",L{row}+1,IF(J{row}="외상3일",L{row}+3,IF(J{row}="외상7일",L{row}+7,IF(J{row}="외상15일",L{row}+15,IF(J{row}="외상30일",L{row}+30,IF(J{row}="외상60일",L{row}+60,IF(J{row}="외상90일",L{row}+90,IF(J{row}="어음60일",L{row}+60,IF(J{row}="어음90일",L{row}+90,""))))))))))))')
        ws.cell(row=row, column=19,
                value=f'=IF(I{row}="",0,IF(O{row}="O",0,IF(OR(M{row}="O",P{row}="O",R{row}="사업자B"),I{row},0)))')
        ws.cell(row=row, column=20,
                value=f'=IF(OR(S{row}=0,N{row}=""),"",TODAY()-N{row})')

    set_widths(ws, [12, 14, 11, 18, 10, 14, 8, 8, 14, 11, 10, 12, 11, 12, 11, 11, 10, 12, 14, 10, 22])

    # 포맷
    for row in range(2, 102):
        ws.cell(row=row, column=8).number_format = '#,##0.##;[Red]-#,##0.##;-'
        for col in [9, 19]:
            ws.cell(row=row, column=col).number_format = '#,##0;[Red]-#,##0;-'
        ws.cell(row=row, column=1).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=12).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=14).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=20).number_format = '0"일";[Red]+0"일";"-"'
        for col in [1, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 17, 18, 21]:
            cell = ws.cell(row=row, column=col)
            if cell.font == DEFAULT_FONT or cell.font == Font():
                cell.font = INPUT_FONT
        for col in [2, 3, 14, 19, 20]:
            ws.cell(row=row, column=col).font = FORMULA_FONT

    # 드롭다운
    dv_pay = DataValidation(type='list', formula1='"즉시,현금,외상1일,외상3일,외상7일,외상15일,외상30일,외상60일,외상90일,어음60일,어음90일"', allow_blank=True)
    dv_pay.add('J2:J102')
    ws.add_data_validation(dv_pay)
    dv_ent = DataValidation(type='list', formula1='"법인,사업자"', allow_blank=True)
    dv_ent.add('K2:K102')
    ws.add_data_validation(dv_ent)
    dv_ox = DataValidation(type='list', formula1='"O,X"', allow_blank=True)
    for col in ['M2:M102', 'O2:O102', 'P2:P102', 'Q2:Q102']:
        dv_ox.add(col)
    ws.add_data_validation(dv_ox)
    dv_account = DataValidation(type='list', formula1='"법인A,사업자A,사업자B,현금,어음"', allow_blank=True)
    dv_account.add('R2:R102')
    ws.add_data_validation(dv_account)

    # 조건부 서식 — 상태 강조
    ws.conditional_formatting.add('C2:C102',
        FormulaRule(formula=['$C2="주문"'], fill=YELLOW_FILL,
                    font=Font(name=FONT_NAME, size=10, bold=True, color='8B6914')))
    ws.conditional_formatting.add('C2:C102',
        FormulaRule(formula=['$C2="납품완료"'], fill=LIGHT_BLUE_FILL,
                    font=Font(name=FONT_NAME, size=10, bold=True, color='2C5F8A')))
    ws.conditional_formatting.add('C2:C102',
        FormulaRule(formula=['$C2="수금완료"'], fill=GREEN_FILL,
                    font=Font(name=FONT_NAME, size=10, bold=True, color='006600')))
    ws.conditional_formatting.add('C2:C102',
        FormulaRule(formula=['$C2="연체"'], fill=RED_FILL,
                    font=Font(name=FONT_NAME, size=10, bold=True, color='CC0000')))
    ws.conditional_formatting.add('A2:U102',
        FormulaRule(formula=['$C2="수금완료"'], fill=PatternFill('solid', start_color='F0FFF0')))
    ws.conditional_formatting.add('A2:U102',
        FormulaRule(formula=['$C2="연체"'], fill=PatternFill('solid', start_color='FFF5F5')))
    ws.conditional_formatting.add('R2:R102',
        FormulaRule(formula=['$R2="사업자B"'],
                    fill=PatternFill('solid', start_color='4A4A4A'),
                    font=Font(name=FONT_NAME, size=10, bold=True, color='FFFFFF')))
    ws.conditional_formatting.add('T2:T102',
        FormulaRule(formula=['AND(ISNUMBER($T2),$T2>=8)'], fill=RED_FILL))
    ws.conditional_formatting.add('T2:T102',
        FormulaRule(formula=['AND(ISNUMBER($T2),$T2>=1,$T2<=7)'], fill=YELLOW_FILL))
    for col_letter in ['M', 'P', 'Q']:
        ws.conditional_formatting.add(f'{col_letter}2:{col_letter}102',
            FormulaRule(formula=[f'${col_letter}2="O"'], fill=GREEN_FILL))

    ws.freeze_panes = 'D2'


# ============================================================
# 시트 2: 매입
# ============================================================
def _build_sheet_purchases(wb, with_samples):
    ws = wb.create_sheet('2.매입')

    headers = [
        '기록일', '매입ID', '상태',
        '매입처', '품목', '규격', '단위', '수량', '금액(원)',
        '결제방식', '사업자',
        '입고일자', '거래명세서수취',
        '결제예정일', '결제완료', '세금계산서수취',
        '세금계산서번호',
        '출금통장',
        '미지급금(원)', '지연일수', '메모'
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    if with_samples:
        samples = [
            [date(2026, 5, 3), None, None, '경기메탈', '강관', '50각', '개', 500, 2250000,
             '외상30일', '법인', None, '', None, '', '', '', '법인A', None, None, '발주, 입고 대기'],
            [date(2026, 5, 2), None, None, 'OO철강', '철근', 'D16', '톤', 5, 4250000,
             '즉시', '법인', date(2026, 5, 2), 'O', None, '', 'O', '20260502-001', '법인A',
             None, None, '입고완료, 결제예정'],
            [date(2026, 5, 4), None, None, 'OO철강', '각파이프', '50×50×2.0t', '개', 200, 900000,
             '즉시', '법인', date(2026, 5, 4), 'O', None, 'O', 'O', '20260504-002', '법인A',
             None, None, '결제완료'],
            [date(2026, 5, 5), None, None, '개인 김씨', '고철', '잉여철근', '톤', 1.2, 300000,
             '현금', '사업자', date(2026, 5, 5), '', None, '', '', '', '사업자B',
             None, None, '무자료 고철 매입'],
        ]
        for s in samples:
            ws.append(s)

    for row in range(2, 102):
        ws.cell(row=row, column=2,
                value=f'=IF($A{row}="","",TEXT($A{row},"YYYYMMDD")&"-"&TEXT(COUNTIF($A$2:$A{row},$A{row}),"000"))')
        ws.cell(row=row, column=3,
                value=f'=IF($D{row}="","",IF($O{row}="O","결제완료",IF(OR($M{row}="O",$P{row}="O",$R{row}="사업자B"),IF(AND($N{row}<>"",$N{row}<TODAY()),"결제연체","입고완료"),"발주")))')
        ws.cell(row=row, column=14,
                value=f'=IF(L{row}="","",IF(J{row}="즉시",L{row},IF(J{row}="현금",L{row},IF(J{row}="외상1일",L{row}+1,IF(J{row}="외상3일",L{row}+3,IF(J{row}="외상7일",L{row}+7,IF(J{row}="외상15일",L{row}+15,IF(J{row}="외상30일",L{row}+30,IF(J{row}="외상60일",L{row}+60,IF(J{row}="외상90일",L{row}+90,IF(J{row}="어음60일",L{row}+60,IF(J{row}="어음90일",L{row}+90,""))))))))))))')
        ws.cell(row=row, column=19,
                value=f'=IF(I{row}="",0,IF(O{row}="O",0,IF(OR(M{row}="O",P{row}="O",R{row}="사업자B"),I{row},0)))')
        ws.cell(row=row, column=20,
                value=f'=IF(OR(S{row}=0,N{row}=""),"",TODAY()-N{row})')

    set_widths(ws, [12, 14, 11, 18, 10, 14, 8, 8, 14, 11, 10, 12, 11, 12, 11, 11, 18, 12, 14, 10, 22])

    for row in range(2, 102):
        ws.cell(row=row, column=8).number_format = '#,##0.##;[Red]-#,##0.##;-'
        for col in [9, 19]:
            ws.cell(row=row, column=col).number_format = '#,##0;[Red]-#,##0;-'
        ws.cell(row=row, column=1).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=12).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=14).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=20).number_format = '0"일";[Red]+0"일";"-"'
        for col in [1, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 17, 18, 21]:
            cell = ws.cell(row=row, column=col)
            if cell.font == DEFAULT_FONT or cell.font == Font():
                cell.font = INPUT_FONT
        for col in [2, 3, 14, 19, 20]:
            ws.cell(row=row, column=col).font = FORMULA_FONT

    dv_pay2 = DataValidation(type='list', formula1='"즉시,현금,외상1일,외상3일,외상7일,외상15일,외상30일,외상60일,외상90일,어음60일,어음90일"', allow_blank=True)
    dv_pay2.add('J2:J102')
    ws.add_data_validation(dv_pay2)
    dv_ent2 = DataValidation(type='list', formula1='"법인,사업자"', allow_blank=True)
    dv_ent2.add('K2:K102')
    ws.add_data_validation(dv_ent2)
    dv_ox2 = DataValidation(type='list', formula1='"O,X"', allow_blank=True)
    for col in ['M2:M102', 'O2:O102', 'P2:P102']:
        dv_ox2.add(col)
    ws.add_data_validation(dv_ox2)
    dv_acc2 = DataValidation(type='list', formula1='"법인A,사업자A,사업자B,현금,어음"', allow_blank=True)
    dv_acc2.add('R2:R102')
    ws.add_data_validation(dv_acc2)

    ws.conditional_formatting.add('C2:C102',
        FormulaRule(formula=['$C2="발주"'], fill=YELLOW_FILL,
                    font=Font(name=FONT_NAME, size=10, bold=True, color='8B6914')))
    ws.conditional_formatting.add('C2:C102',
        FormulaRule(formula=['$C2="입고완료"'], fill=LIGHT_BLUE_FILL,
                    font=Font(name=FONT_NAME, size=10, bold=True, color='2C5F8A')))
    ws.conditional_formatting.add('C2:C102',
        FormulaRule(formula=['$C2="결제완료"'], fill=GREEN_FILL,
                    font=Font(name=FONT_NAME, size=10, bold=True, color='006600')))
    ws.conditional_formatting.add('C2:C102',
        FormulaRule(formula=['$C2="결제연체"'], fill=RED_FILL,
                    font=Font(name=FONT_NAME, size=10, bold=True, color='CC0000')))
    ws.conditional_formatting.add('A2:U102',
        FormulaRule(formula=['$C2="결제완료"'], fill=PatternFill('solid', start_color='F0FFF0')))
    ws.conditional_formatting.add('A2:U102',
        FormulaRule(formula=['$C2="결제연체"'], fill=PatternFill('solid', start_color='FFF5F5')))
    ws.conditional_formatting.add('R2:R102',
        FormulaRule(formula=['$R2="사업자B"'],
                    fill=PatternFill('solid', start_color='4A4A4A'),
                    font=Font(name=FONT_NAME, size=10, bold=True, color='FFFFFF')))
    ws.conditional_formatting.add('T2:T102',
        FormulaRule(formula=['AND(ISNUMBER($T2),$T2>=8)'], fill=RED_FILL))
    ws.conditional_formatting.add('T2:T102',
        FormulaRule(formula=['AND(ISNUMBER($T2),$T2>=1,$T2<=7)'], fill=YELLOW_FILL))
    for col_letter in ['M', 'P']:
        ws.conditional_formatting.add(f'{col_letter}2:{col_letter}102',
            FormulaRule(formula=[f'${col_letter}2="O"'], fill=GREEN_FILL))

    ws.freeze_panes = 'D2'


# ============================================================
# 시트 3: 재고
# ============================================================
def _build_sheet_inventory(wb, with_samples):
    ws = wb.create_sheet('3.재고')

    headers = [
        '품목', '규격', '단위', '사업자', '현재수량', '안전재고',
        '평균단가(원)', '재고가치(원)', '마지막입고일', '마지막출고일',
        '회전상태', '메모'
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    if with_samples:
        samples = [
            ['철근', 'D10', '톤', '법인', 5.5, 3.0, 480000, None, date(2026, 5, 1), date(2026, 5, 4), None, '예시'],
            ['철근', 'D13', '톤', '법인', 8.2, 5.0, 500000, None, date(2026, 5, 2), date(2026, 5, 3), None, ''],
            ['철근', 'D16', '톤', '법인', 12.0, 5.0, 850000, None, date(2026, 5, 2), '', None, ''],
            ['강관', '50×50×2.0t', '개', '법인', 350, 200, 3000, None, date(2026, 5, 3), date(2026, 5, 4), None, ''],
            ['철근', 'D13(중고폴리싱)', '톤', '사업자', 2.5, 1.0, 380000, None, date(2026, 5, 4), '', None, '폴리싱 재고'],
        ]
        for s in samples:
            ws.append(s)

    for row in range(2, 102):
        ws.cell(row=row, column=8, value=f'=IF(OR(E{row}="",G{row}=""),"",E{row}*G{row})')
        ws.cell(row=row, column=11,
                value=f'=IF(E{row}="","",IF(E{row}<F{row},"⚠ 안전재고 미달",IF(E{row}<F{row}*1.5,"주의","정상")))')

    set_widths(ws, [10, 18, 8, 12, 12, 12, 14, 16, 14, 14, 18, 22])

    for row in range(2, 102):
        for col in [5, 6]:
            ws.cell(row=row, column=col).number_format = '#,##0.##;[Red]-#,##0.##;-'
        for col in [7, 8]:
            ws.cell(row=row, column=col).number_format = '#,##0;[Red]-#,##0;-'
        ws.cell(row=row, column=9).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=10).number_format = 'yyyy-mm-dd'
        for col in [1, 2, 3, 4, 5, 6, 7, 9, 10, 12]:
            ws.cell(row=row, column=col).font = INPUT_FONT
        for col in [8, 11]:
            ws.cell(row=row, column=col).font = FORMULA_FONT

    dv_unit = DataValidation(type='list', formula1='"톤,kg,개,m,장,세트"', allow_blank=True)
    dv_unit.add('C2:C102')
    ws.add_data_validation(dv_unit)
    dv_ent3 = DataValidation(type='list', formula1='"법인,사업자"', allow_blank=True)
    dv_ent3.add('D2:D102')
    ws.add_data_validation(dv_ent3)

    ws.conditional_formatting.add('E2:E102',
        FormulaRule(formula=['AND($E2<>"",$F2<>"",$E2<$F2)'], fill=RED_FILL))
    ws.conditional_formatting.add('E2:E102',
        FormulaRule(formula=['AND($E2<>"",$F2<>"",$E2>=$F2,$E2<$F2*1.5)'], fill=YELLOW_FILL))

    ws.freeze_panes = 'A2'


# ============================================================
# 시트: 5.거래처
# ============================================================
def _build_sheet_customers(wb, with_samples):
    ws = wb.create_sheet('5.거래처')

    headers = [
        '구분', '거래처명', '대표자', '연락처', '사업자번호',
        '주소/현장', '거래시작일', '결제조건', '선호사업자',
        '미수금(원)', '누적매출(원)', '메모'
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    if with_samples:
        samples = [
            ['매출', 'OO건설', '김사장', '010-1234-5678', '123-45-67890', '경주시 황성동',
             date(2026, 5, 3), '외상30일', '법인', None, None, '주력 거래처'],
            ['매출', '박사장', '박OO', '010-2345-6789', '', '울산 중구',
             date(2026, 5, 4), '즉시', '법인', None, None, '소규모 시공'],
            ['매출', '소형시공', '이OO', '010-3456-7890', '', '포항 북구',
             date(2026, 5, 5), '현금', '사업자', None, None, '무자료 거래'],
            ['매입', 'OO철강', '최사장', '010-4567-8901', '234-56-78901', '경기도 안산',
             date(2026, 5, 2), '즉시', '법인', None, None, '주력 매입처'],
            ['매입', '경기메탈', '이사장', '010-5678-9012', '345-67-89012', '경기도 화성',
             date(2026, 5, 3), '외상30일', '법인', None, None, '강관 전문'],
            ['매출', '김씨건축', None, None, '', '',
             date(2026, 5, 6), '현금', '사업자', None, None, '무자료 거래 (B통장)'],
            ['매입', '개인 김씨', None, None, '', '',
             date(2026, 5, 5), '현금', '사업자', None, None, '무자료 고철 매입 (B통장)'],
        ]
        for s in samples:
            ws.append(s)

    for row in range(2, 52):
        ws.cell(row=row, column=10,
                value=f'=IF(A{row}="매출",IFERROR(SUMIFS(\'1.매출\'!S:S,\'1.매출\'!D:D,B{row}),0),'
                      f'IF(A{row}="매입",IFERROR(SUMIFS(\'2.매입\'!S:S,\'2.매입\'!D:D,B{row}),0),""))')
        ws.cell(row=row, column=11,
                value=f'=IF(A{row}="매출",IFERROR(SUMIFS(\'1.매출\'!I:I,\'1.매출\'!D:D,B{row}),0),'
                      f'IF(A{row}="매입",IFERROR(SUMIFS(\'2.매입\'!I:I,\'2.매입\'!D:D,B{row}),0),""))')

    set_widths(ws, [8, 18, 10, 16, 16, 22, 12, 12, 12, 14, 14, 22])

    for row in range(2, 52):
        for col in [10, 11]:
            ws.cell(row=row, column=col).number_format = '#,##0;[Red]-#,##0;-'
        ws.cell(row=row, column=7).number_format = 'yyyy-mm-dd'
        for col in [1, 2, 3, 4, 5, 6, 7, 8, 9, 12]:
            ws.cell(row=row, column=col).font = INPUT_FONT
        for col in [10, 11]:
            ws.cell(row=row, column=col).font = FORMULA_FONT

    dv_type = DataValidation(type='list', formula1='"매출,매입"', allow_blank=True)
    dv_type.add('A2:A52')
    ws.add_data_validation(dv_type)
    dv_pay4 = DataValidation(type='list', formula1='"즉시,현금,외상30일,외상60일,외상90일"', allow_blank=True)
    dv_pay4.add('H2:H52')
    ws.add_data_validation(dv_pay4)
    dv_ent4 = DataValidation(type='list', formula1='"법인,사업자"', allow_blank=True)
    dv_ent4.add('I2:I52')
    ws.add_data_validation(dv_ent4)

    ws.freeze_panes = 'A2'


# ============================================================
# 시트: 6.통장
# ============================================================
def _build_sheet_bank(wb, with_samples):
    ws = wb.create_sheet('6.통장')

    headers = [
        '일자', '사업자', '통장', '적요', '입금(원)', '출금(원)',
        '누적잔고(원)', '분류', '매칭ID', '메모'
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    if with_samples:
        start_balances = [
            [date(2026, 5, 1), '법인', '법인A', '시작 잔고', 0, 0, 100000000, '시작', '', '본인 자본금 1.5억 + 운영자금'],
            [date(2026, 5, 1), '사업자', '사업자A', '시작 잔고', 0, 0, 5000000, '시작', '', '사업자 운영자금'],
            [date(2026, 5, 1), '사업자', '사업자B', '시작 잔고', 0, 0, 0, '시작', '', '히든 통장 — 무자료 거래 입금'],
        ]
        for r in start_balances:
            ws.append(r)
        samples = [
            [date(2026, 5, 3), '법인', '법인A', 'OO건설 매출 입금', 1250000, 0, None, '매출입금', '20260503-001', '매출 1번'],
            [date(2026, 5, 3), '법인', '법인A', '식대', 0, 35000, None, '운영비', '', '직원 식대'],
            [date(2026, 5, 4), '법인', '법인A', '박사장 매출 입금', 600000, 0, None, '매출입금', '20260501-001', ''],
            [date(2026, 5, 4), '법인', '법인A', '운반비 지출', 0, 50000, None, '운영비', '', 'A현장 운반'],
            [date(2026, 5, 6), '사업자', '사업자B', '김씨건축 무자료 입금', 250000, 0, None, '매출입금(B통장)', '20260506-001', '히든 통장'],
        ]
        for s in samples:
            ws.append(s)

    for row in range(2, 152):
        ws.cell(row=row, column=7,
                value=f'=IF(C{row}="","",SUMIFS($E$2:E{row},$C$2:C{row},C{row})-SUMIFS($F$2:F{row},$C$2:C{row},C{row}))')

    set_widths(ws, [12, 10, 12, 25, 14, 14, 16, 18, 14, 22])

    for row in range(2, 152):
        for col in [5, 6, 7]:
            ws.cell(row=row, column=col).number_format = '#,##0;[Red]-#,##0;-'
        ws.cell(row=row, column=1).number_format = 'yyyy-mm-dd'
        for col in [1, 2, 3, 4, 5, 6, 8, 9, 10]:
            ws.cell(row=row, column=col).font = INPUT_FONT
        ws.cell(row=row, column=7).font = FORMULA_FONT

    dv_ent5 = DataValidation(type='list', formula1='"법인,사업자"', allow_blank=True)
    dv_ent5.add('B2:B152')
    ws.add_data_validation(dv_ent5)
    dv_acc = DataValidation(type='list', formula1='"법인A,사업자A,사업자B,현금"', allow_blank=True)
    dv_acc.add('C2:C152')
    ws.add_data_validation(dv_acc)
    dv_cat = DataValidation(type='list', formula1='"시작,매출입금,매출입금(B통장),매입출금,임원보수,임대료,법인카드결제,운영비,세금,차용,기타"', allow_blank=True)
    dv_cat.add('H2:H152')
    ws.add_data_validation(dv_cat)

    ws.conditional_formatting.add('E2:E152',
        FormulaRule(formula=['AND($E2<>"",$E2>0)'], fill=LIGHT_BLUE_FILL))
    ws.conditional_formatting.add('A2:J152',
        FormulaRule(formula=['$C2="사업자B"'],
                    fill=PatternFill('solid', start_color='3A3A3A'),
                    font=Font(name=FONT_NAME, size=10, bold=True, color='FFFFFF')))

    # 통장별 잔고 요약 (오른쪽 L~O)
    ws['L1'] = '통장별 현재 잔고'
    ws['L1'].font = SUBTITLE_FONT
    ws.merge_cells('L1:N1')
    ws['L2'] = '통장'
    ws['M2'] = '입금합계'
    ws['N2'] = '출금합계'
    ws['O2'] = '현재잔고'
    for col in ['L', 'M', 'N', 'O']:
        cell = ws[f'{col}2']
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    accounts_summary = ['법인A', '사업자A', '사업자B', '현금']
    for i, acc in enumerate(accounts_summary):
        r = 3 + i
        ws[f'L{r}'] = acc
        ws[f'M{r}'] = f'=SUMIFS(E:E,C:C,L{r})'
        ws[f'N{r}'] = f'=SUMIFS(F:F,C:C,L{r})'
        ws[f'O{r}'] = f'=M{r}-N{r}'
        for col in ['M', 'N', 'O']:
            ws[f'{col}{r}'].number_format = '#,##0;[Red]-#,##0;-'
            ws[f'{col}{r}'].font = FORMULA_FONT
        if acc == '사업자B':
            ws[f'L{r}'].font = Font(name=FONT_NAME, size=10, bold=True, color='4A4A4A')
        else:
            ws[f'L{r}'].font = INPUT_FONT

    ws['L7'] = '합계'
    ws['L7'].font = Font(name=FONT_NAME, size=11, bold=True)
    ws['M7'] = '=SUM(M3:M6)'
    ws['N7'] = '=SUM(N3:N6)'
    ws['O7'] = '=SUM(O3:O6)'
    for col in ['M', 'N', 'O']:
        ws[f'{col}7'].font = Font(name=FONT_NAME, size=11, bold=True, color='2C5F8A')
        ws[f'{col}7'].number_format = '#,##0"원";[Red]-#,##0"원";-'

    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 14
    ws.column_dimensions['N'].width = 14
    ws.column_dimensions['O'].width = 16

    ws.conditional_formatting.add('L3:O6',
        FormulaRule(formula=['$L3="사업자B"'],
                    fill=PatternFill('solid', start_color='E8E8E8')))

    ws.freeze_panes = 'A2'


# ============================================================
# 시트: 현황판 (수식만, 데이터는 없음)
# ============================================================
def _build_sheet_dashboard(wb):
    ws = wb.create_sheet('현황판', 0)

    ws['A1'] = 'SH철강 현황판'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 30

    ws['A2'] = '*숫자가 자동 업데이트됩니다. 매출·매입·통장 시트에 데이터 입력 시 즉시 반영.'
    ws['A2'].font = Font(name=FONT_NAME, size=9, italic=True, color='666666')
    ws.merge_cells('A2:F2')

    ws['A4'] = '이번 달 매출'; ws['A4'].font = SUBTITLE_FONT
    ws['A5'] = '법인'
    ws['B5'] = '=SUMIFS(\'1.매출\'!I:I,\'1.매출\'!A:A,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),\'1.매출\'!A:A,"<"&DATE(YEAR(TODAY()),MONTH(TODAY())+1,1),\'1.매출\'!K:K,"법인")'
    ws['A6'] = '사업자'
    ws['B6'] = '=SUMIFS(\'1.매출\'!I:I,\'1.매출\'!A:A,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),\'1.매출\'!A:A,"<"&DATE(YEAR(TODAY()),MONTH(TODAY())+1,1),\'1.매출\'!K:K,"사업자")'
    ws['A7'] = '합계'; ws['A7'].font = Font(name=FONT_NAME, size=10, bold=True)
    ws['B7'] = '=B5+B6'; ws['B7'].font = Font(name=FONT_NAME, size=10, bold=True)

    ws['D4'] = '이번 달 매입'; ws['D4'].font = SUBTITLE_FONT
    ws['D5'] = '법인'
    ws['E5'] = '=SUMIFS(\'2.매입\'!I:I,\'2.매입\'!A:A,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),\'2.매입\'!A:A,"<"&DATE(YEAR(TODAY()),MONTH(TODAY())+1,1),\'2.매입\'!K:K,"법인")'
    ws['D6'] = '사업자'
    ws['E6'] = '=SUMIFS(\'2.매입\'!I:I,\'2.매입\'!A:A,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),\'2.매입\'!A:A,"<"&DATE(YEAR(TODAY()),MONTH(TODAY())+1,1),\'2.매입\'!K:K,"사업자")'
    ws['D7'] = '합계'; ws['D7'].font = Font(name=FONT_NAME, size=10, bold=True)
    ws['E7'] = '=E5+E6'; ws['E7'].font = Font(name=FONT_NAME, size=10, bold=True)

    ws['A9'] = '이번 달 추정 영업이익'; ws['A9'].font = SUBTITLE_FONT
    ws['A10'] = '매출 - 매입'
    ws['B10'] = '=B7-E7'; ws['B10'].font = Font(name=FONT_NAME, size=11, bold=True, color='2C5F8A')

    ws['A12'] = '미수금 현황 (등급별)'; ws['A12'].font = SUBTITLE_FONT
    ws['A13'] = '정상 (예정일 미도래)'
    ws['B13'] = '=IFERROR(SUMPRODUCT((\'1.매출\'!N2:N102>=TODAY())*(\'1.매출\'!N2:N102<>"")*(\'1.매출\'!O2:O102<>"O")*(\'1.매출\'!S2:S102)),0)'
    ws['A14'] = '단기 연체 (1~7일)'
    ws['B14'] = '=IFERROR(SUMPRODUCT((\'1.매출\'!N2:N102<TODAY())*(\'1.매출\'!N2:N102>=TODAY()-7)*(\'1.매출\'!N2:N102<>"")*(\'1.매출\'!O2:O102<>"O")*(\'1.매출\'!S2:S102)),0)'
    ws['A15'] = '중기 연체 (8~30일)'
    ws['B15'] = '=IFERROR(SUMPRODUCT((\'1.매출\'!N2:N102<TODAY()-7)*(\'1.매출\'!N2:N102>=TODAY()-30)*(\'1.매출\'!N2:N102<>"")*(\'1.매출\'!O2:O102<>"O")*(\'1.매출\'!S2:S102)),0)'
    ws['A16'] = '장기 연체 (31일 이상)'
    ws['B16'] = '=IFERROR(SUMPRODUCT((\'1.매출\'!N2:N102<TODAY()-30)*(\'1.매출\'!N2:N102<>"")*(\'1.매출\'!O2:O102<>"O")*(\'1.매출\'!S2:S102)),0)'
    ws['A17'] = '전체 미수금 합계'; ws['A17'].font = Font(name=FONT_NAME, size=10, bold=True)
    ws['B17'] = '=B13+B14+B15+B16'; ws['B17'].font = Font(name=FONT_NAME, size=10, bold=True)

    ws['D12'] = '미지급금 현황'; ws['D12'].font = SUBTITLE_FONT
    ws['D13'] = '전체 미지급금 합계'
    ws['E13'] = '=SUM(\'2.매입\'!S:S)'
    ws['D14'] = '연체 미지급금 (지난 결제예정일)'
    ws['E14'] = '=IFERROR(SUMPRODUCT((\'2.매입\'!N2:N102<TODAY())*(\'2.매입\'!N2:N102<>"")*(\'2.매입\'!O2:O102<>"O")*(\'2.매입\'!S2:S102)),0)'

    ws['D19'] = '통장별 잔고'; ws['D19'].font = SUBTITLE_FONT
    ws['D20'] = '법인A'
    ws['E20'] = "=SUMIFS('6.통장'!E:E,'6.통장'!C:C,\"법인A\")-SUMIFS('6.통장'!F:F,'6.통장'!C:C,\"법인A\")"
    ws['D21'] = '사업자A'
    ws['E21'] = "=SUMIFS('6.통장'!E:E,'6.통장'!C:C,\"사업자A\")-SUMIFS('6.통장'!F:F,'6.통장'!C:C,\"사업자A\")"
    ws['F20'] = '사업자B (히든)'; ws['F20'].font = Font(name=FONT_NAME, size=10, bold=True, color='4A4A4A')
    ws['G20'] = "=SUMIFS('6.통장'!E:E,'6.통장'!C:C,\"사업자B\")-SUMIFS('6.통장'!F:F,'6.통장'!C:C,\"사업자B\")"
    ws['F21'] = '현금'
    ws['G21'] = "=SUMIFS('6.통장'!E:E,'6.통장'!C:C,\"현금\")-SUMIFS('6.통장'!F:F,'6.통장'!C:C,\"현금\")"

    ws['A22'] = '재고 알림'; ws['A22'].font = SUBTITLE_FONT
    ws['A23'] = '안전재고 미달 품목 수'
    ws['B23'] = '=SUMPRODUCT((\'3.재고\'!E2:E102<>"")*(\'3.재고\'!F2:F102<>"")*(\'3.재고\'!E2:E102<\'3.재고\'!F2:F102))'

    ws['D22'] = '정기업무 알림'; ws['D22'].font = SUBTITLE_FONT
    ws['D23'] = '이번 주 임박 (7일 내)'
    ws['E23'] = '=SUMPRODUCT((\'7.정기업무\'!E5:E100<>"")*(\'7.정기업무\'!E5:E100>=TODAY())*(\'7.정기업무\'!E5:E100<=TODAY()+7)*(\'7.정기업무\'!G5:G100<>"완료"))'
    ws['D24'] = '지연 (미완료)'
    ws['E24'] = '=SUMPRODUCT((\'7.정기업무\'!E5:E100<>"")*(\'7.정기업무\'!E5:E100<TODAY())*(\'7.정기업무\'!G5:G100<>"완료"))'
    ws['D25'] = '이번 달 예상 비용'
    ws['E25'] = '=SUMPRODUCT((\'7.정기업무\'!E5:E100<>"")*(\'7.정기업무\'!E5:E100>=DATE(YEAR(TODAY()),MONTH(TODAY()),1))*(\'7.정기업무\'!E5:E100<DATE(YEAR(TODAY()),MONTH(TODAY())+1,1))*(\'7.정기업무\'!H5:H100))'

    for cell in ['A4', 'D4', 'A9', 'A12', 'D12', 'D19', 'A22', 'D22']:
        ws[cell].fill = LIGHT_BLUE_FILL

    set_widths(ws, [22, 18, 4, 22, 18, 18])

    won_cells = ['B5', 'B6', 'B7', 'B10', 'B13', 'B14', 'B15', 'B16', 'B17',
                 'E5', 'E6', 'E7', 'E13', 'E14', 'E20', 'E21', 'G20', 'G21', 'E25']
    for c in won_cells:
        ws[c].number_format = '#,##0"원";[Red]-#,##0"원";-'

    ws['G20'].font = Font(name=FONT_NAME, size=11, bold=True, color='4A4A4A')
    ws['B23'].number_format = '0"개"'
    ws['E23'].number_format = '0"건"'
    ws['E24'].number_format = '0"건";[Red]0"건"'
    ws['B14'].font = Font(name=FONT_NAME, size=10, color='B45F06')
    ws['B15'].font = Font(name=FONT_NAME, size=10, bold=True, color='D4700E')
    ws['B16'].font = Font(name=FONT_NAME, size=10, bold=True, color='CC0000')
    ws['E23'].font = Font(name=FONT_NAME, size=11, bold=True, color='2C5F8A')
    ws['E24'].font = Font(name=FONT_NAME, size=11, bold=True, color='D4700E')
    ws['B10'].number_format = '#,##0"원";[Red]-#,##0"원";-'

    ws['A27'] = '운영 룰'; ws['A27'].font = SUBTITLE_FONT
    notes = [
        '※ 미수금 입금 완료 시 매출 시트의 수금완료(O열)을 "O"로 변경',
        '※ 매입 결제 완료 시 매입 시트의 결제완료(O열)을 "O"로 변경',
        '※ 통장 입출금 발생 시 통장 시트에 추가 입력',
        '※ 정기업무 완료 시 7.정기업무 시트의 상태를 "완료"로 변경 + 다음 실행일 갱신',
        '※ 미수 등급: 정상 / 단기(1~7일) / 중기(8~30일) / 장기(31일+)',
        '※ 거래처별 미수 상세는 8.미수관리 시트 참조',
    ]
    for i, note in enumerate(notes):
        r = 28 + i
        ws.cell(row=r, column=1, value=note)
        ws.cell(row=r, column=1).font = Font(name=FONT_NAME, size=9, color='666666')
        ws.merge_cells(f'A{r}:F{r}')


# ============================================================
# 시트: 0.개선아이디어
# ============================================================
def _build_sheet_ideas(wb, with_samples):
    ws = wb.create_sheet('0.개선아이디어', 1)

    ws['A1'] = '시스템 개선 아이디어'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 30
    ws['A2'] = '*5월 운영하면서 발견한 개선 아이디어 누적. 6월 시스템 개발 시 to-do 리스트로 활용.'
    ws['A2'].font = Font(name=FONT_NAME, size=9, italic=True, color='666666')
    ws.merge_cells('A2:F2')

    headers = ['일자', '영역', '발견한 문제·아이디어', '시스템 반영 우선순위', '상태', '메모']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[4].height = 32

    if with_samples:
        samples = [
            [date(2026, 5, 3), '매출', '거래처가 "납품 시각"을 자주 물어봄. 매출 시트에 컬럼 추가 필요', '중', '대기', '예시'],
            [date(2026, 5, 5), '통장', '거래처명 "OO건설" vs "OO 건설" 표기 차이로 자동 매칭 실패', '상', '대기', '예시: 거래처명 표준화 필요'],
            [date(2026, 5, 8), '재고', '폴리싱 작업 시간·작업자 기록할 곳 없음', '하', '대기', '예시'],
        ]
        for idx, sample in enumerate(samples):
            for col_num, value in enumerate(sample, 1):
                ws.cell(row=5 + idx, column=col_num, value=value)

    set_widths(ws, [12, 12, 50, 18, 12, 30])
    for row in range(5, 55):
        ws.cell(row=row, column=1).number_format = 'yyyy-mm-dd'
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            if cell.font.name != FONT_NAME:
                cell.font = INPUT_FONT
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    dv_area = DataValidation(type='list', formula1='"매출,매입,재고,거래처,통장,현황판,자동화,알림,UI,세무,기타"', allow_blank=True)
    dv_area.add('B5:B55')
    ws.add_data_validation(dv_area)
    dv_pri = DataValidation(type='list', formula1='"상,중,하"', allow_blank=True)
    dv_pri.add('D5:D55')
    ws.add_data_validation(dv_pri)
    dv_status = DataValidation(type='list', formula1='"대기,검토중,반영예정,반영완료,보류"', allow_blank=True)
    dv_status.add('E5:E55')
    ws.add_data_validation(dv_status)

    ws.conditional_formatting.add('D5:D55',
        FormulaRule(formula=['$D5="상"'], fill=RED_FILL))
    ws.conditional_formatting.add('D5:D55',
        FormulaRule(formula=['$D5="중"'], fill=YELLOW_FILL))
    ws.conditional_formatting.add('E5:E55',
        FormulaRule(formula=['$E5="반영완료"'], fill=GREEN_FILL))

    ws.freeze_panes = 'A5'


# ============================================================
# 시트: 7.정기업무
# ============================================================
def _build_sheet_recurring_tasks(wb, with_samples):
    ws = wb.create_sheet('7.정기업무')

    ws['A1'] = '정기업무 관리'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:I1')
    ws.row_dimensions[1].height = 30
    ws['A2'] = '*차량 검사·세금 신고·시설 점검·정기 결제 등 잊으면 안 되는 일을 한 곳에 관리'
    ws['A2'].font = Font(name=FONT_NAME, size=9, italic=True, color='666666')
    ws.merge_cells('A2:I2')

    headers = ['분류', '업무명', '주기', '마지막 실행일', '다음 실행일',
               '담당자', '상태', '예상비용(원)', '메모']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[4].height = 32

    if with_samples:
        preset_tasks = [
            ['차량', '5톤 크레인트럭 자동차 정기검사', '매년', date(2025, 8, 10), date(2026, 8, 15), '본인', '예정', 100000, '신경주 자동차 검사장'],
            ['차량', '크레인 안전검사 (산업안전)', '매년', date(2025, 9, 5), date(2026, 9, 5), '본인', '예정', 200000, '한국안전인증원'],
            ['차량', '엔진오일 교환', '매3개월', date(2026, 3, 15), date(2026, 6, 15), '직원', '예정', 80000, '단골 정비소'],
            ['차량', '타이어 점검', '매6개월', date(2025, 12, 1), date(2026, 6, 1), '직원', '예정', 50000, ''],
            ['차량', '자동차보험 갱신', '매년', date(2025, 7, 1), date(2026, 7, 1), '본인', '예정', 1500000, ''],
            ['차량', '자동차세 납부 (1기)', '매년 6월', date(2025, 6, 30), date(2026, 6, 30), '본인', '예정', 300000, ''],
            ['차량', '자동차세 납부 (2기)', '매년 12월', date(2025, 12, 31), date(2026, 12, 31), '본인', '예정', 300000, ''],
            ['세무', '부가세 신고 (1분기)', '매년 4월 25일', date(2026, 4, 25), date(2027, 4, 25), '본인', '완료', 0, '세무사 위임'],
            ['세무', '부가세 신고 (2분기)', '매년 7월 25일', None, date(2026, 7, 25), '본인', '예정', 0, '세무사 위임'],
            ['세무', '부가세 신고 (3분기)', '매년 10월 25일', None, date(2026, 10, 25), '본인', '예정', 0, '세무사 위임'],
            ['세무', '부가세 신고 (4분기)', '매년 1월 25일', None, date(2027, 1, 25), '본인', '예정', 0, '세무사 위임'],
            ['세무', '종합소득세 신고', '매년 5월', date(2025, 5, 31), date(2026, 5, 31), '본인', '예정', 0, '세무사 위임'],
            ['세무', '원천세 신고', '매월 10일', date(2026, 4, 10), date(2026, 5, 10), '본인', '예정', 0, '직원 급여 원천세'],
            ['세무', '4대보험 신고', '매월', date(2026, 4, 15), date(2026, 5, 15), '본인', '예정', 0, ''],
            ['시설', '공장 임대료 지급', '매월 1일', date(2026, 4, 1), date(2026, 5, 1), '본인', '예정', 8000000, '친구에게 자동이체'],
            ['시설', '소방시설 점검', '매년', date(2025, 11, 10), date(2026, 11, 10), '본인', '예정', 200000, ''],
            ['시설', '전기 안전점검', '매년', date(2025, 10, 20), date(2026, 10, 20), '본인', '예정', 150000, ''],
            ['인력', '직원 급여 지급', '매월 25일', date(2026, 4, 25), date(2026, 5, 25), '본인', '예정', 3500000, ''],
            ['인력', '직원 건강검진', '매년', date(2025, 8, 1), date(2026, 8, 1), '직원', '예정', 0, '국민건강보험공단 무료'],
            ['관계', '명절 거래처 인사 (설)', '매년 2월', date(2026, 2, 5), date(2027, 2, 5), '친구', '예정', 1000000, '주요 거래처'],
            ['관계', '명절 거래처 인사 (추석)', '매년 9월', date(2025, 9, 20), date(2026, 9, 20), '친구', '예정', 1000000, '주요 거래처'],
            ['관계', '거래처 신용 점검', '매분기', date(2026, 4, 1), date(2026, 7, 1), '본인', '예정', 0, '주요 거래처 신용정보 확인'],
            ['재무', '월말 결산', '매월 말일', date(2026, 4, 30), date(2026, 5, 31), '본인', '예정', 0, ''],
            ['재무', '재고 실사', '매월 말일', date(2026, 4, 30), date(2026, 5, 31), '직원', '예정', 0, ''],
            ['재무', '분기 결산 검토', '매분기', date(2026, 3, 31), date(2026, 6, 30), '본인', '예정', 0, '세무사와 검토'],
        ]
        for idx, task in enumerate(preset_tasks):
            for col_num, value in enumerate(task, 1):
                ws.cell(row=5 + idx, column=col_num, value=value)

    set_widths(ws, [10, 30, 14, 14, 14, 10, 10, 14, 25])

    for row in range(5, 100):
        ws.cell(row=row, column=4).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=5).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=8).number_format = '#,##0;[Red]-#,##0;-'
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            if cell.font.name != FONT_NAME:
                cell.font = INPUT_FONT
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    dv_cat_t = DataValidation(type='list', formula1='"차량,세무,시설,인력,관계,재무,기타"', allow_blank=True)
    dv_cat_t.add('A5:A100')
    ws.add_data_validation(dv_cat_t)
    dv_period = DataValidation(type='list',
        formula1='"매일,매주,매월,매월 1일,매월 10일,매월 25일,매월 말일,매분기,매2개월,매3개월,매6개월,매년,매년 1월,매년 2월,매년 5월,매년 6월,매년 7월,매년 9월,매년 10월,매년 11월,매년 12월,일회성"',
        allow_blank=True)
    dv_period.add('C5:C100')
    ws.add_data_validation(dv_period)
    dv_owner = DataValidation(type='list', formula1='"본인,친구,직원,세무사,외주"', allow_blank=True)
    dv_owner.add('F5:F100')
    ws.add_data_validation(dv_owner)
    dv_status_t = DataValidation(type='list', formula1='"예정,진행중,완료,지연,보류"', allow_blank=True)
    dv_status_t.add('G5:G100')
    ws.add_data_validation(dv_status_t)

    ws.conditional_formatting.add('E5:E100',
        FormulaRule(formula=['AND($E5<>"",$E5<TODAY(),$G5<>"완료")'], fill=RED_FILL))
    ws.conditional_formatting.add('E5:E100',
        FormulaRule(formula=['AND($E5<>"",$E5-TODAY()<=7,$E5-TODAY()>=0,$G5<>"완료")'], fill=YELLOW_FILL))
    ws.conditional_formatting.add('G5:G100',
        FormulaRule(formula=['$G5="완료"'], fill=GREEN_FILL))
    ws.conditional_formatting.add('G5:G100',
        FormulaRule(formula=['$G5="지연"'], fill=RED_FILL))

    ws.freeze_panes = 'A5'


# ============================================================
# 시트: 8.미수관리
# ============================================================
def _build_sheet_receivables(wb, with_samples):
    ws = wb.create_sheet('8.미수관리')

    ws['A1'] = '미수관리 — 거래처별 집계'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 30
    ws['A2'] = '*거래처명을 입력하면 매출 시트에서 자동 집계. 미수금 0원이면 행 숨기시면 됨.'
    ws['A2'].font = Font(name=FONT_NAME, size=9, italic=True, color='666666')
    ws.merge_cells('A2:H2')

    headers = ['거래처명', '미수금 합계(원)', '미수 건수',
               '정상 미수(원)', '단기 연체(원)', '중기 연체(원)', '장기 연체(원)',
               '최악 등급', '마지막 독촉일', '메모']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[4].height = 32

    if with_samples:
        preset_customers = ['OO건설', '박사장', '소형시공']
        for idx, customer in enumerate(preset_customers):
            ws.cell(row=5 + idx, column=1, value=customer)

    for row in range(5, 55):
        ws.cell(row=row, column=2,
            value=f"=IF(A{row}=\"\",\"\",IFERROR(SUMIFS(\'1.매출\'!S:S,\'1.매출\'!D:D,A{row}),0))")
        ws.cell(row=row, column=3,
            value=f"=IF(A{row}=\"\",\"\",IFERROR(SUMPRODUCT((\'1.매출\'!D2:D102=A{row})*(\'1.매출\'!S2:S102>0)),0))")
        ws.cell(row=row, column=4,
            value=f"=IF(OR(A{row}=\"\",B{row}=0),\"\",IFERROR(SUMPRODUCT((\'1.매출\'!D2:D102=A{row})*ISNUMBER(\'1.매출\'!T2:T102)*(\'1.매출\'!T2:T102<=0)*\'1.매출\'!S2:S102),0))")
        ws.cell(row=row, column=5,
            value=f"=IF(OR(A{row}=\"\",B{row}=0),\"\",IFERROR(SUMPRODUCT((\'1.매출\'!D2:D102=A{row})*ISNUMBER(\'1.매출\'!T2:T102)*(\'1.매출\'!T2:T102>0)*(\'1.매출\'!T2:T102<=7)*\'1.매출\'!S2:S102),0))")
        ws.cell(row=row, column=6,
            value=f"=IF(OR(A{row}=\"\",B{row}=0),\"\",IFERROR(SUMPRODUCT((\'1.매출\'!D2:D102=A{row})*ISNUMBER(\'1.매출\'!T2:T102)*(\'1.매출\'!T2:T102>7)*(\'1.매출\'!T2:T102<=30)*\'1.매출\'!S2:S102),0))")
        ws.cell(row=row, column=7,
            value=f"=IF(OR(A{row}=\"\",B{row}=0),\"\",IFERROR(SUMPRODUCT((\'1.매출\'!D2:D102=A{row})*ISNUMBER(\'1.매출\'!T2:T102)*(\'1.매출\'!T2:T102>30)*\'1.매출\'!S2:S102),0))")
        ws.cell(row=row, column=8,
            value=f'=IF(OR(A{row}="",B{row}=0),"",IF(G{row}>0,"장기",IF(F{row}>0,"중기",IF(E{row}>0,"단기",IF(D{row}>0,"정상","")))))')

    set_widths(ws, [18, 14, 10, 14, 14, 14, 14, 11, 14, 22])

    for row in range(5, 55):
        for col in [2, 4, 5, 6, 7]:
            ws.cell(row=row, column=col).number_format = '#,##0;[Red]-#,##0;-'
        ws.cell(row=row, column=3).number_format = '0"건"'
        ws.cell(row=row, column=9).number_format = 'yyyy-mm-dd'
        for col in [1, 9, 10]:
            cell = ws.cell(row=row, column=col)
            if cell.font.name != FONT_NAME:
                cell.font = INPUT_FONT
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        for col in [2, 3, 4, 5, 6, 7, 8]:
            ws.cell(row=row, column=col).font = FORMULA_FONT

    ws.conditional_formatting.add('H5:H55',
        FormulaRule(formula=['$H5="장기"'], fill=RED_FILL))
    ws.conditional_formatting.add('H5:H55',
        FormulaRule(formula=['$H5="중기"'], fill=YELLOW_FILL))
    ws.conditional_formatting.add('H5:H55',
        FormulaRule(formula=['$H5="정상"'], fill=GREEN_FILL))
    ws.conditional_formatting.add('G5:G55',
        FormulaRule(formula=['$G5>0'], fill=RED_FILL))
    ws.conditional_formatting.add('F5:F55',
        FormulaRule(formula=['$F5>0'], fill=YELLOW_FILL))

    ws['A57'] = '합계'; ws['A57'].font = Font(name=FONT_NAME, size=11, bold=True)
    for col_letter, formula in [('B', '=SUM(B5:B55)'), ('C', '=SUM(C5:C55)'),
                                 ('D', '=SUM(D5:D55)'), ('E', '=SUM(E5:E55)'),
                                 ('F', '=SUM(F5:F55)'), ('G', '=SUM(G5:G55)')]:
        cell = ws[f'{col_letter}57']
        cell.value = formula
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color='2C5F8A')
        if col_letter == 'C':
            cell.number_format = '0"건"'
        else:
            cell.number_format = '#,##0"원"'

    ws.freeze_panes = 'A5'


# ============================================================
# 시트: 4.영수증
# ============================================================
def _build_sheet_receipts(wb, with_samples):
    ws = wb.create_sheet('4.영수증')

    ws['A1'] = '비품·영수증 관리'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 30
    ws['A2'] = '*비품 구매·식대·판촉물 등 소액 지출 영수증을 한 곳에 보관. 시스템화 시 expenses 테이블로 import.'
    ws['A2'].font = Font(name=FONT_NAME, size=9, italic=True, color='666666')
    ws.merge_cells('A2:H2')

    headers = ['날짜', '영수증ID', '항목', '분류', '금액(원)', '사업자', '영수증첨부', '메모']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[4].height = 32

    if with_samples:
        samples = [
            [date(2026, 5, 3), None, '거래처용 음료수 1박스', '판촉물', 35000, '법인', '', 'OO건설 미팅용'],
            [date(2026, 5, 4), None, '직원 점심 식대', '식대', 28000, '법인', '', '직원과 함께'],
            [date(2026, 5, 5), None, '명함 200매 인쇄', '판촉물', 50000, '법인', '', ''],
            [date(2026, 5, 6), None, '주유 (5톤트럭)', '차량유지', 80000, '법인', '', 'A-1234호'],
            [date(2026, 5, 7), None, '운반비 (현장)', '운반비', 30000, 'B계좌', '', '히든 자금 사용'],
        ]
        for idx, sample in enumerate(samples):
            for col_num, value in enumerate(sample, 1):
                ws.cell(row=5 + idx, column=col_num, value=value)

    for row in range(5, 205):
        ws.cell(row=row, column=2,
                value=f'=IF($A{row}="","","REC-"&TEXT($A{row},"YYYYMMDD")&"-"&TEXT(COUNTIF($A$5:$A{row},$A{row}),"000"))')

    set_widths(ws, [12, 18, 25, 12, 13, 11, 30, 25])

    for row in range(5, 205):
        ws.cell(row=row, column=1).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=5).number_format = '#,##0;[Red]-#,##0;-'
        for col in [1, 3, 4, 5, 6, 7, 8]:
            cell = ws.cell(row=row, column=col)
            if cell.font.name != FONT_NAME:
                cell.font = INPUT_FONT
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(row=row, column=2).font = FORMULA_FONT

    dv_cat_r = DataValidation(type='list', formula1='"식대,차량유지,판촉물,사무용품,운반비,통신비,접대비,경조사,기타"', allow_blank=True)
    dv_cat_r.add('D5:D205')
    ws.add_data_validation(dv_cat_r)
    dv_ent_r = DataValidation(type='list', formula1='"법인,사업자,B계좌"', allow_blank=True)
    dv_ent_r.add('F5:F205')
    ws.add_data_validation(dv_ent_r)

    ws.conditional_formatting.add('F5:F205',
        FormulaRule(formula=['$F5="B계좌"'],
                    fill=PatternFill('solid', start_color='4A4A4A'),
                    font=Font(name=FONT_NAME, size=10, bold=True, color='FFFFFF')))
    ws.conditional_formatting.add('G5:G205',
        FormulaRule(formula=['$G5<>""'],
                    fill=PatternFill('solid', start_color='F0FFF0')))

    # 분류별·사업자별 합계 (오른쪽 J~L)
    ws['J1'] = '이번 달 분류별 합계'
    ws['J1'].font = SUBTITLE_FONT
    ws.merge_cells('J1:L1')

    ws['J2'] = '분류'; ws['K2'] = '건수'; ws['L2'] = '합계(원)'
    for col in ['J', 'K', 'L']:
        cell = ws[f'{col}2']
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    categories_r = ['식대', '차량유지', '판촉물', '사무용품', '운반비', '통신비', '접대비', '경조사', '기타']
    for i, cat in enumerate(categories_r):
        r = 3 + i
        ws[f'J{r}'] = cat
        ws[f'K{r}'] = f'=COUNTIFS(D5:D205,J{r},A5:A205,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),A5:A205,"<"&DATE(YEAR(TODAY()),MONTH(TODAY())+1,1))'
        ws[f'L{r}'] = f'=SUMIFS(E5:E205,D5:D205,J{r},A5:A205,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),A5:A205,"<"&DATE(YEAR(TODAY()),MONTH(TODAY())+1,1))'
        ws[f'J{r}'].font = INPUT_FONT
        ws[f'K{r}'].font = FORMULA_FONT
        ws[f'L{r}'].font = FORMULA_FONT
        ws[f'K{r}'].number_format = '0"건"'
        ws[f'L{r}'].number_format = '#,##0;[Red]-#,##0;-'

    total_r = 3 + len(categories_r)
    ws[f'J{total_r}'] = '합계'
    ws[f'J{total_r}'].font = Font(name=FONT_NAME, size=11, bold=True)
    ws[f'K{total_r}'] = f'=SUM(K3:K{total_r - 1})'
    ws[f'L{total_r}'] = f'=SUM(L3:L{total_r - 1})'
    ws[f'K{total_r}'].font = Font(name=FONT_NAME, size=11, bold=True, color='2C5F8A')
    ws[f'L{total_r}'].font = Font(name=FONT_NAME, size=11, bold=True, color='2C5F8A')
    ws[f'K{total_r}'].number_format = '0"건"'
    ws[f'L{total_r}'].number_format = '#,##0"원";[Red]-#,##0"원";-'

    ws['J15'] = '이번 달 사업자별 합계'
    ws['J15'].font = SUBTITLE_FONT
    ws.merge_cells('J15:L15')
    ws['J16'] = '사업자'; ws['K16'] = '건수'; ws['L16'] = '합계(원)'
    for col in ['J', 'K', 'L']:
        cell = ws[f'{col}16']
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    entities_r = ['법인', '사업자', 'B계좌']
    for i, ent in enumerate(entities_r):
        r = 17 + i
        ws[f'J{r}'] = ent
        ws[f'K{r}'] = f'=COUNTIFS(F5:F205,J{r},A5:A205,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),A5:A205,"<"&DATE(YEAR(TODAY()),MONTH(TODAY())+1,1))'
        ws[f'L{r}'] = f'=SUMIFS(E5:E205,F5:F205,J{r},A5:A205,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),A5:A205,"<"&DATE(YEAR(TODAY()),MONTH(TODAY())+1,1))'
        if ent == 'B계좌':
            ws[f'J{r}'].font = Font(name=FONT_NAME, size=10, bold=True, color='4A4A4A')
        else:
            ws[f'J{r}'].font = INPUT_FONT
        ws[f'K{r}'].font = FORMULA_FONT
        ws[f'L{r}'].font = FORMULA_FONT
        ws[f'K{r}'].number_format = '0"건"'
        ws[f'L{r}'].number_format = '#,##0;[Red]-#,##0;-'

    total_e = 17 + len(entities_r)
    ws[f'J{total_e}'] = '합계'
    ws[f'J{total_e}'].font = Font(name=FONT_NAME, size=11, bold=True)
    ws[f'K{total_e}'] = f'=SUM(K17:K{total_e - 1})'
    ws[f'L{total_e}'] = f'=SUM(L17:L{total_e - 1})'
    ws[f'K{total_e}'].font = Font(name=FONT_NAME, size=11, bold=True, color='2C5F8A')
    ws[f'L{total_e}'].font = Font(name=FONT_NAME, size=11, bold=True, color='2C5F8A')
    ws[f'K{total_e}'].number_format = '0"건"'
    ws[f'L{total_e}'].number_format = '#,##0"원";[Red]-#,##0"원";-'

    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 8
    ws.column_dimensions['L'].width = 14

    ws.freeze_panes = 'A5'


# ============================================================
# 진입점
# ============================================================
if __name__ == '__main__':
    import sys
    args = sys.argv[1:]
    output = None
    with_samples = True
    for a in args:
        if a == '--no-samples':
            with_samples = False
        else:
            output = a
    if output is None:
        output = 'workbook/output/SL철강_5월_운영시트.xlsx'

    build_workbook(output, with_samples=with_samples)
    print(f'✓ {output} (samples={with_samples})')
